import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.common.keys import Keys
import sys
import os
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import warnings
warnings.filterwarnings("ignore")

# -------- OPEN LOGIN PAGE --------
options = uc.ChromeOptions()

options.binary_location = "/usr/bin/google-chrome"

options.add_argument("--headless=new")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--disable-gpu")

driver = uc.Chrome(
    options=options,
    driver_executable_path="/usr/bin/chromedriver",
    use_subprocess=True
)

wait = WebDriverWait(driver, 20)

# ================= OPEN WEBSITE =================
driver.get("https://www.cma-cgm.com/")
driver.maximize_window()

time.sleep(5)

# ================= OPEN MY CMA =================
my_cma = wait.until(
    EC.element_to_be_clickable(
        (
            By.XPATH,
            "//a[contains(@class,'mycma')]"
        )
    )
)

driver.execute_script(
    "arguments[0].click();",
    my_cma
)

time.sleep(3)

# ================= LOGIN BUTTON =================
login_btn = wait.until(
    EC.element_to_be_clickable(
        (
            By.XPATH,
            "//button[contains(text(),'Log In')]"
        )
    )
)

driver.execute_script(
    "arguments[0].click();",
    login_btn
)

time.sleep(3)

# ================= EMAIL =================
email = wait.until(
    EC.presence_of_element_located(
        (
            By.ID,
            "login-email"
        )
    )
)

email.clear()

email.send_keys(
    "Ocean@tanuimpex.com"
)

# ================= PASSWORD =================
password = wait.until(
    EC.presence_of_element_located(
        (
            By.ID,
            "login-password"
        )
    )
)

password.clear()

password.send_keys(
    "Mtanu2025#"
)

# ================= FINAL LOGIN =================
final_login = wait.until(
    EC.element_to_be_clickable(
        (
            By.CSS_SELECTOR,
            "button.o-button.primary"
        )
    )
)

driver.execute_script(
    "arguments[0].click();",
    final_login
)

print("CMA CGM Login Success")

time.sleep(10)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))              
file_path = os.path.join(BASE_DIR, "global", "TrackerDashBoard-7-March.xlsm")
print(BASE_DIR)
print(file_path)    


df = pd.read_excel(file_path, sheet_name="NotSailed")
df.columns = df.columns.str.strip()

# workbook load
wb = load_workbook(file_path, keep_vba=True)
ws = wb["NotSailed"]

cmacgm_df = df[
    df["Shipping line"].str.lower().str.contains("cma cgm", na=False) &
    df["SSL Booking"].notna() &
    (df["SSL Booking"].astype(str).str.strip() != "")
]

print("Total CMA CGM Bookings:", len(cmacgm_df))

def format_date(value):
    try:
        value = str(value).strip()

        if value == "-" or value == "":
            return "-"

        return datetime.strptime(value, "%m/%d/%Y").strftime("%d-%m-%Y")

    except:
        return "-"

for idx, row in cmacgm_df.iterrows():

    driver.get("https://www.cma-cgm.com/ebusiness/tracking")
    time.sleep(2)

    booking = str(row["SSL Booking"]).strip()

    print("\nOpening:", booking)

    try:

        # ================= BOOKING INPUT =================
        reference_input = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//input[@id='Reference']"))
        )

        reference_input.clear()
        reference_input.send_keys(booking)

        search_btn = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "button[id='btnTracking']"))
        )

        driver.execute_script("arguments[0].click();", search_btn)

        # ================= OPEN BOOKING =================
        try:
            booking_link1 = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located(
                    (By.XPATH, "//span[@class='value-Info']/a")
                )
            )

            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});",
                booking_link1
            )

            driver.execute_script(
                "window.location.href = arguments[0].href;",
                booking_link1
            )

        except:

            booking_link2 = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located(
                    (By.XPATH, "//ul[@class='resume-filter']/li/a")
                )
            )

            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});",
                booking_link2
            )

            driver.execute_script(
                "window.location.href = arguments[0].href;",
                booking_link2
            )

        # ================= TRANSPORT PLAN TAB =================
        transport_plan = wait.until(
            EC.element_to_be_clickable((
                By.XPATH,
                "//a[contains(@data-href,'transport-plan')]"
            ))
        )

        driver.execute_script("arguments[0].click();", transport_plan)

        time.sleep(2)

        print("Booking:", booking)

        # ================= EXPORT SAIL DATE =================
        export_sail_date = wait.until(
            EC.presence_of_element_located((
                By.XPATH,
                "//dt[contains(text(),'Export Sail Date POL')]/following-sibling::dd"
            ))
        ).text.split()[0]

        # ================= IMPORT ETA DATE =================
        import_eta_date = wait.until(
            EC.presence_of_element_located((
                By.XPATH,
                "//dt[contains(text(),'Import ETA Date POD')]/following-sibling::dd"
            ))
        ).text.split()[0]

        # ================= PORT CUT OFF =================
        port_cutoff = wait.until(
            EC.presence_of_element_located((
                By.XPATH,
                "//dt[contains(text(),'Port Cut-off')]/following-sibling::dd"
            ))
        ).text.strip()[0]

        # ================= FORMAT DATE =================
        etd = format_date(export_sail_date)
        eta = format_date(import_eta_date)
        cut_off = format_date(port_cutoff)

        print("ETD :", etd)
        print("ETA :", eta)
        print("Port Cut-off :", cut_off)

        # ================= EXCEL UPDATE =================
        excel_row = idx + 2

        ws[f"M{excel_row}"] = etd
        ws[f"P{excel_row}"] = eta
        ws[f"L{excel_row}"] = cut_off

    except Exception as e:
        print("Error:", e)


# ================= SAVE FILE =================
wb.save(file_path)

print("\nData saved successfully")

driver.quit()