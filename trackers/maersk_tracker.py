import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
import pandas as pd
from selenium.common.exceptions import TimeoutException
import time

# =========================================================
# LOOP BOOKINGS
# =========================================================
from openpyxl import load_workbook
import warnings
warnings.filterwarnings("ignore")

# -------- DRIVER SETUP --------
options = uc.ChromeOptions()

options.binary_location = "/usr/bin/google-chrome"

options.add_argument("--headless=new")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--disable-gpu")

driver = uc.Chrome(
    options=options,
    driver_executable_path="/usr/bin/chromedriver",
    use_subprocess=True
)
wait = WebDriverWait(driver, 30)

# -------- OPEN SITE --------
driver.get("https://www.maersk.com/")
time.sleep(3)

# -------- COOKIE ACCEPT --------
try:
    driver.find_element(By.CSS_SELECTOR, "button[data-test='coi-allow-all-button']").click()
    time.sleep(2)
except:
    pass

# -------- LOGIN --------
driver.find_element(By.CSS_SELECTOR, "a[class='ign-header__primary__actions__links__link ign-button-primary   ign-track']").click()
time.sleep(10) 

for el in driver.find_elements(By.CSS_SELECTOR, "mc-input"):
    shadow = driver.execute_script("return arguments[0].shadowRoot", el)
    inp = shadow.find_element(By.CSS_SELECTOR, "input")
    
    if inp.get_attribute("id") == "mc-input-username":
        inp.send_keys("Tanuimpex")
        
    elif inp.get_attribute("id") == "mc-input-password":
        inp.send_keys("MyMaersk123")

time.sleep(5)

driver.execute_script("""
document.querySelector("mc-button#login-submit-button")
.shadowRoot.querySelector("button").click();
""")

time.sleep(8)

# -------- GO TO TRACKING --------
driver.get("https://www.maersk.com/tracking/")
time.sleep(5)

# -------- EXCEL READ --------
import os
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))              
file_path = os.path.join(BASE_DIR, "global", "TrackerDashBoard-7-March.xlsm")
print(BASE_DIR)
print(file_path)    

df = pd.read_excel(file_path, sheet_name="NotSailed")
df.columns = df.columns.str.strip()

wb = load_workbook(file_path, keep_vba=True)
ws = wb["NotSailed"]

# -------- DATE FORMAT --------
def format_eta(value):
    try:
        return datetime.strptime(value, "%d %b %Y %H:%M").strftime("%d-%m-%Y")
    except:
        try:
            return datetime.strptime(value, "%d %b %Y").strftime("%d-%m-%Y")
        except:
            return ""

# -------- FILTER MAERSK --------
maersk_df = df[
    df["Shipping line"].str.lower().str.contains("maersk", na=False) &
    df["SSL Booking"].notna() &
    (df["SSL Booking"].astype(str).str.strip() != "")
]

print("Total Maersk Bookings:", len(maersk_df))

# -------- ETA FUNCTION --------
def latest_event():

    # ⏳ Wait for page load
    time.sleep(5)

    for _ in range(3):  # retry

        eta_value = driver.execute_script("""
        let blocks = document.querySelectorAll('mc-text-and-icon');

        for (let b of blocks){
            let root = b.shadowRoot;
            if (!root) continue;

            let text = root.innerText;

            if (text && text.includes('Estimated arrival')){
                let match = text.match(/\\d{2} \\w{3} \\d{4}( \\d{2}:\\d{2})?/);
                if (match){
                    return match[0];
                }
            }
        }
        return '';
        """)

        if eta_value:
            return eta_value

        #fallback: पूरे page से ढूंढो
        eta_value = driver.execute_script("""
        let txt = document.body.innerText;
        let match = txt.match(/\\d{2} \\w{3} \\d{4}( \\d{2}:\\d{2})?/);
        return match ? match[0] : '';
        """)

        if eta_value:
            return eta_value

        time.sleep(2)

    return ""
    
def get_eta():

    # ⏳ wait for render
    time.sleep(5)

    for _ in range(3):

        eta_value = driver.execute_script("""
        let blocks = document.querySelectorAll('mc-text-and-icon');

        for (let b of blocks){
            let root = b.shadowRoot;
            if (!root) continue;

            let labelBox = root.querySelector('.labels');

            if (!labelBox) continue;

            let text = labelBox.innerText;

            if (text && text.toLowerCase().includes('estimated arrival')){

                let lines = text.split('\\n');

                if (lines.length > 1){
                    return lines[1].trim();  // 👉 second line = date
                }
            }
        }
        return '';
        """)

        print("DEBUG ETA:", eta_value)

        if eta_value:
            return eta_value

        time.sleep(2)

    return ""


# =========================================================
# LOOP
# =========================================================

for idx, row in maersk_df.iterrows():

    booking = str(row["SSL Booking"]).strip()

    print("\n Opening:", booking)

    try:

        # =====================================================
        # OPEN PAGE
        # =====================================================
        driver.get(
            f"https://www.maersk.com/shipment-details/{booking}/summary?app=SD&searchTerm={booking}"
        )

        # =====================================================
        # WAIT PAGE LOAD
        # =====================================================
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located(
                (
                    By.CSS_SELECTOR,
                    '[data-test="shipment-summary-from-to-date"]'
                )
            )
        )

        time.sleep(3)

        # =====================================================
        # ETD
        # =====================================================
        etd = ""

        try:
            etd_text = driver.find_element(
                By.CSS_SELECTOR,
                '[data-test="shipment-summary-from-to-date"]'
            ).text

            etd_raw = etd_text.split("\n")[-1].strip()

            etd = datetime.strptime(
                etd_raw,
                "%d %b %Y"
            ).strftime("%d-%m-%Y")

        except:
            pass

        # =====================================================
        # ETA
        # =====================================================
        eta = ""

        try:
            eta_text = driver.find_element(
                By.CSS_SELECTOR,
                '[data-test="shipment-summary-destination-date"]'
            ).text

            eta_raw = eta_text.split("\n")[-1].strip()

            eta = datetime.strptime(
                eta_raw,
                "%d %b %Y"
            ).strftime("%d-%m-%Y")

        except:
            pass

        print("ETD :", etd)
        print("ETA :", eta)

        # =====================================================
        # TASK VARIABLES
        # =====================================================
        pick_empty_container = ""
        submit_shipping_instruction = ""
        submit_vgm = ""

        # =====================================================
        # CLICK VIEW ALL TASKS
        # =====================================================
        try:

            view_tasks_btn = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable(
                    (
                        By.CSS_SELECTOR,
                        '[data-test="tasks-documents-view-all-tasks"]'
                    )
                )
            )

            driver.execute_script(
                "arguments[0].scrollIntoView(true);",
                view_tasks_btn
            )

            time.sleep(1)

            driver.execute_script(
                "arguments[0].click();",
                view_tasks_btn
            )

            print("View All Tasks clicked")

        except Exception as e:
            print("View All Tasks Error:", e)

        # =====================================================
        # WAIT MODAL LOAD
        # =====================================================
        try:

            WebDriverWait(driver, 30).until(
                EC.visibility_of_element_located(
                    (
                        By.CSS_SELECTOR,
                        'mc-modal[data-test="tasks-modal"][open]'
                    )
                )
            )

            time.sleep(3)

        except TimeoutException:
            print("Modal not opened")

        # =====================================================
        # GET TASK CARDS
        # =====================================================
        task_cards = driver.find_elements(
            By.XPATH,
            '//mc-modal[@data-test="tasks-modal" and @open]//div[@data-test="task-card"]'
        )

        print(f"Total Tasks Found : {len(task_cards)}")

        # =====================================================
        # LOOP TASKS
        # =====================================================
        for task in task_cards:

            task_name = ""
            due_date = ""
           
            try:

                task_name = task.find_element(
                    By.XPATH,
                    './/*[@data-test="task-name" or @data-test="task-name-VGM" or @data-test="task-name-SIAMS"]'
                ).text.strip()

            except:
                pass

            try:

                due_date = task.find_element(
                    By.CSS_SELECTOR,
                    '[data-test="tasks-due-date"]'
                ).text.strip()

            except:
                pass

            if due_date == "Deadline not available":
                due_date = "Update"

            if "Pick empty container" in task_name:
                pick_empty_container = due_date

            elif "Submit shipping instruction" in task_name:
                submit_shipping_instruction = due_date

            elif "Submit VGM" in task_name:
                submit_vgm = due_date
       
        print("Pick empty container :", pick_empty_container)
        print("Submit shipping instruction :", submit_shipping_instruction)
        print("Submit VGM :", submit_vgm)

        # =====================================================
        # EXCEL UPDATE
        # =====================================================
        # Pick empty container = erd
        # Submit shipping instruction = Doc_Cut_off
        # Submit VGM = VGM_Cut_off
        excel_row = idx + 2

        ws[f"M{excel_row}"] = etd
        ws[f"P{excel_row}"] = eta

        ws[f"I{excel_row}"] = pick_empty_container
        ws[f"J{excel_row}"] = submit_shipping_instruction
        ws[f"K{excel_row}"] = submit_vgm

        # =====================================================
        # CLOSE MODAL
        # =====================================================
        try:

            close_btn = driver.find_element(
                By.CSS_SELECTOR,
                '[data-test="tasks-modal-close-button"]'
            )

            driver.execute_script(
                "arguments[0].click();",
                close_btn
            )

            time.sleep(1)

        except:
            pass

    except Exception as e:

        print("Error:", booking, e)

        continue


# -------- SAVE --------
wb.save(file_path)

print("Excel updated successfully")

driver.quit()