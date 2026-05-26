from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time
import undetected_chromedriver as uc
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
import pandas as pd    
import re
import warnings
import sys
import os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
warnings.filterwarnings("ignore")

# ================= DRIVER =================
options = uc.ChromeOptions()
options.add_argument("--start-maximized")
options.add_argument("--disable-blink-features=AutomationControlled")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
profile_path = os.path.join(BASE_DIR, "chrome", "hapag")
options.add_argument(f"--user-data-dir={profile_path}")


driver = uc.Chrome(
    options=options,
    version_main=148,
    use_subprocess=True
)

# ================= OPEN SITE =================

driver.get("https://www.hapag-lloyd.com/en/home.html")
driver.maximize_window()
time.sleep(10)
wait = WebDriverWait(driver, 30)
# ================= ACCEPT COOKIES =================

try:

    wait.until(
        EC.element_to_be_clickable(
            (By.ID, "accept-recommended-btn-handler")
        )
    ).click()

    print("cookies accepted")

except:
    pass

time.sleep(5)

# ================= LOGIN CHECK =================

body = driver.find_element(
    By.TAG_NAME,
    "body"
).text.lower()

if (
    "gitanjali" in body
    or "my hapag" in body
    or "sign out" in body
    or "logout" in body
    or "tanu impex" in body
):

    print("already logged in")

else:

    print("login required") #div.hal-navigation-top-login-web

    try:

        # ================= CLICK LOGIN =================

        login_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div.hal-navigation-top-login-web"))).click()  

        print("clicked login")
        time.sleep(10)

        # ================= USERNAME =================

        username = wait.until(
            EC.visibility_of_element_located(
                (By.ID, "signInName")
            )
        )

        username.clear()

        username.send_keys(
            "shipping@tanuimpex.com"
        )

        print("username entered")

        # ================= PASSWORD =================

        password = wait.until(
            EC.visibility_of_element_located(
                (By.ID, "password")
            )
        )

        password.clear()

        password.send_keys(
            "MyHapag123"
        )

        print("password entered")

        time.sleep(2)

        # ================= LOGIN BUTTON =================

        driver.find_element(
            By.ID,
            "next"
        ).click()

        print("login submitted")

        time.sleep(20)

        # ================= FINAL CHECK =================

        body = driver.find_element(
            By.TAG_NAME,
            "body"
        ).text.lower()

        if (
            "gitanjali" in body
            or "my hapag" in body
            or "sign out" in body
        ):

            print("login completed")

        else:

            print("login may have failed")

    except Exception as e:

        print("login failed")
        print(e)

# ================= READY =================

print("ready for scraping")
time.sleep(20)

driver.execute_script("window.location.href='/solutions/navigator/#/?language=en'")
time.sleep(20)
import os
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))              
file_path = os.path.join(BASE_DIR, "global", "TrackerDashBoard-7-March.xlsm")
print(BASE_DIR)
print(file_path)    

df = pd.read_excel(file_path, sheet_name="NotSailed")
df.columns = df.columns.str.strip()

# workbook load
wb = load_workbook(file_path, keep_vba=True)
ws = wb["NotSailed"]

# ================= DATE FORMAT =================
def format_date(value):
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").strftime("%d-%m-%Y")
    except:
        return ""

# ================= LOOP =================
for i, row in df.iterrows():

    shipping = str(row["Shipping line"]).strip().lower()
    booking_data = str(row["SSL Booking"]).strip()

    # skip empty
    if booking_data == "" or booking_data.lower() == "nan":
        continue

    # only hapag
    if "hapag" not in shipping:
        continue

    # ================= MULTIPLE BOOKING =================
    # comma / slash / newline split
    bookings = re.split(r"[,/\n]+", booking_data)

    for booking in bookings:

        booking = booking.strip()

        if booking == "":
            continue

        print("=" * 50)
        print("Searching:", booking)

        try:

            # ================= SEARCH BOX =================
            search_box = wait.until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, "input[placeholder='e.g. 12345678']")
                )
            )

            # clear old value
            search_box.clear()
            time.sleep(1)

            search_box.send_keys(booking)
            search_box.send_keys(Keys.ENTER)

            time.sleep(8)

            # ================= DETAIL ICON =================
            icon = wait.until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, "i.q-icon.text-info-dark")
                )
            )

            driver.execute_script("arguments[0].click();", icon)

            time.sleep(5)

            # ================= EXTRACT DATA =================

            etd_raw = driver.find_element(
                By.XPATH,
                "//div[text()='ETD:']/following-sibling::div"
            ).text

            etd = format_date(etd_raw)

            eta_raw = driver.find_element(
                By.XPATH,
                "//div[text()='ETA:']/following-sibling::div"
            ).text

            eta = format_date(eta_raw)

            vgm_raw = driver.find_element(
                By.XPATH,
                "//div[text()='VGM']/following-sibling::div//span"
            ).text

            vgm_date = vgm_raw.split(" ")[0]
            vgm = format_date(vgm_date)

            doc_raw = driver.find_element(
                By.XPATH,
                "//div[text()='Document closure']/following-sibling::div//span"
            ).text

            doc_date = doc_raw.split(" ")[0]
            doc_closure = format_date(doc_date)

            fcl_raw = driver.find_element(
                By.XPATH,
                "//div[text()='FCL']/following-sibling::div//span"
            ).text

            fcl_date = fcl_raw.split(" ")[0]
            cut_off = format_date(fcl_date)

            # ================= PRINT =================
            print("FCL:", cut_off)
            print("Doc:", doc_closure)
            print("VGM:", vgm)
            print("ETD:", etd)
            print("ETA:", eta)

            # ================= EXCEL UPDATE =================
            excel_row = i + 2

            ws[f"M{excel_row}"] = etd
            ws[f"P{excel_row}"] = eta
            ws[f"K{excel_row}"] = vgm
            ws[f"J{excel_row}"] = doc_closure
            ws[f"L{excel_row}"] = cut_off

            # ================= BACK BUTTON =================
            back_btn = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//button[contains(., 'Back')]")
                )
            )

            driver.execute_script(
                "arguments[0].click();",
                back_btn
            )

            time.sleep(3)

        except Exception as e:

            print("Error:", booking)
            print(e)

            try:
                back_btn = driver.find_element(
                    By.XPATH,
                    "//button[contains(., 'Back')]"
                )

                driver.execute_script(
                    "arguments[0].click();",
                    back_btn
                )

                time.sleep(2)

            except:
                pass

            continue

# ================= SAVE =================
wb.save(file_path)
print("Excel Updated Successfully")
driver.quit()


